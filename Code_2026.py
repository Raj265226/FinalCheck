# tab switch and broken link and read,write excel
import os
from openpyxl.reader.excel import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager
import requests
import pandas as pd

driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
driver.get('https://opensource-demo.orangehrmlive.com/web/index.php/auth/login')
driver.maximize_window()
parent_window = driver.current_window_handle
links = WebDriverWait(driver,10).until(EC.presence_of_all_elements_located((By.XPATH,'//a[@href]')))
data = []
for link in links:
    link.click()
    if not link.get_attribute('href'):                                     # python treats '',0,False,None as a False
        print('no url found')
        data.append([link.get_attribute('href'),'No Url','FAIL'])
    else:
        try:
            response = requests.get(link.get_attribute('href'))
            if response.status_code == 200:
                print('url ->',link.get_attribute('href'),' is working perfectly and status -> ',response.status_code)
                data.append([link.get_attribute('href'),response.status_code,'PASS'])
            else:
                print('url ->',link.get_attribute('href'),' is not working perfectly and status -> ',response.status_code)
                data.append([link.get_attribute('href'),response.status_code,'FAIL'])
        except requests.exceptions.RequestException as e:                 # It handles all request related exceptions
            print('Error',e)
            data.append([link.get_attribute('href'), 'ERROR', 'FAIL'])
df = pd.DataFrame(data,columns=['URL','Status Code','Result'])
if os.path.exists('Link_status.xlsx'):
    print(pd.read_excel('Link_status.xlsx',sheet_name='Link_Status'))     # Read excel file
# df.to_excel('Link_status.xlsx',sheet_name='Link_Status',index=False)   # Create excel file
with pd.ExcelWriter('Link_status.xlsx', engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer: # Append
    df.to_excel(writer, sheet_name='Link_Status', index=False, header=False,startrow=load_workbook('Link_status.xlsx')['Link_Status'].max_row)
windows = driver.window_handles
for window in windows:
    if window != parent_window:
        driver.switch_to.window(window)
        print(driver.title)
        driver.close()
driver.switch_to.window(parent_window)
WebDriverWait(driver, 10).until(EC.number_of_windows_to_be(1))








